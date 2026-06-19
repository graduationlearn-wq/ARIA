"""
Lead importer — parse a CSV / Excel upload into normalised lead records.

Employees upload a contact list (CSV or .xlsx); each row becomes a lead owned
by the importer. Headers are matched flexibly (case/spacing/aliases) so a file
exported from any CRM or a hand-made sheet both work.

CSV is parsed with the stdlib. Excel uses openpyxl, imported lazily so a
deployment without it still handles CSV (and gives a clear message for .xlsx).
"""

import csv
import io
import re

MAX_ROWS = 2000
ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".xls"}

# Canonical field -> accepted header spellings (compared after normalisation:
# lowercased with spaces/underscores/hyphens/dots removed).
HEADER_ALIASES: dict[str, set[str]] = {
    "first_name":       {"name", "firstname", "fullname", "leadname", "contactname", "contact"},
    "email":            {"email", "emailaddress", "mail", "emailid"},
    "phone":            {"phone", "mobile", "phonenumber", "mobilenumber", "contactnumber", "number", "whatsapp"},
    "company_name":     {"company", "companyname", "organisation", "organization", "business", "businessname", "firm"},
    "state":            {"state", "location", "region", "city", "place"},
    "lead_type":        {"type", "leadtype", "businesstype", "category", "segment", "profession"},
    "source":           {"source", "platform", "channel", "leadsource"},
    "team_size":        {"teamsize", "team", "size", "employees", "headcount", "staff"},
    "uses_software":    {"usessoftware", "software", "hassoftware", "currentsoftware"},
    "open_to_platform": {"opentoplatform", "opentonew", "opentoadopting", "interested"},
    "willing_for_demo": {"wantsdemo", "demo", "willingfordemo", "wantdemo", "demointerest"},
    "company_website":  {"website", "url", "companywebsite", "web", "site"},
}

_BOOL_TRUE = {"yes", "y", "true", "1", "t", "✓", "interested"}
_BOOL_FALSE = {"no", "n", "false", "0", "f", "-"}

_LEAD_TYPE_MAP = {
    "broker": "broker", "imf": "imf",
    "agent": "agent", "insuranceagent": "agent", "insurance_agent": "agent",
    "posp": "posp_advisor", "pospadvisor": "posp_advisor", "posp_advisor": "posp_advisor",
    "advisor": "advisor", "invalid": "invalid",
}


def _norm_header(h: str) -> str:
    """Lowercase + drop everything non-alphanumeric (handles long FB question headers)."""
    return re.sub(r"[^a-z0-9]", "", (h or "").strip().lower())


# Keyword rules for the Meta/Facebook Lead Ads export, whose headers are the full
# question text (e.g. "do_you_currently_use_any_software_or_digital_platform_…").
# Each rule: field, and a predicate over the normalised header key. Checked after
# exact-alias matching. Order matters — first match wins.
_KEYWORD_RULES = [
    ("lead_type",        lambda k: "typesofbusiness" in k or "businesstype" in k),
    ("uses_software",    lambda k: "currentlyuse" in k or "usesoftware" in k or ("software" in k and "manage" in k)),
    ("open_to_platform", lambda k: ("opento" in k and ("adopt" in k or "platform" in k or "technology" in k))),
    ("willing_for_demo", lambda k: "willing" in k and ("evaluate" in k or "demo" in k or "trial" in k)),
    ("company_website",  lambda k: "website" in k or "companysite" in k),
]


def _header_to_field(header: str) -> str | None:
    """Map a raw column header to a canonical lead field, or None if unknown."""
    key = _norm_header(header)
    # 1. Exact alias match (the simple, clean headers).
    for field, aliases in HEADER_ALIASES.items():
        if key == field or key in aliases:
            return field
    # 2. Keyword match (the long FB-export question headers).
    for field, pred in _KEYWORD_RULES:
        if pred(key):
            return field
    return None


def _parse_bool(val: str | None):
    if val is None:
        return None
    v = str(val).strip().lower()
    if v in _BOOL_TRUE:
        return True
    if v in _BOOL_FALSE:
        return False
    return None


def _parse_lead_type(val: str | None) -> str:
    if not val:
        return "unknown"
    key = re.sub(r"[\s_\-.]+", "", str(val).strip().lower())
    return _LEAD_TYPE_MAP.get(key, str(val).strip().lower() or "unknown")


def _parse_source(val: str | None) -> str:
    v = (str(val or "").strip().lower())
    if v in ("fb", "facebook"):
        return "facebook"
    if v in ("ig", "insta", "instagram"):
        return "instagram"
    return v or "import"


def _parse_int(val: str | None):
    digits = re.sub(r"[^\d]", "", str(val or ""))
    return int(digits) if digits else None


# ── Parsing ───────────────────────────────────────────────────────────────────

def _parse_csv(data: bytes) -> list[dict]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return [{(k or ""): (v or "") for k, v in row.items()} for row in reader]


def _parse_xlsx(data: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook  # lazy — only needed for Excel
    except ImportError:
        raise ValueError(
            "Excel (.xlsx) support needs the 'openpyxl' package. "
            "Install it, or upload a CSV instead."
        )
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [str(h).strip() if h is not None else "" for h in next(rows)]
    except StopIteration:
        return []
    records = []
    for r in rows:
        if r is None or all(c is None or str(c).strip() == "" for c in r):
            continue  # skip blank rows
        record = {}
        for i, h in enumerate(headers):
            record[h] = "" if i >= len(r) or r[i] is None else str(r[i]).strip()
        records.append(record)
    wb.close()
    return records


def parse_rows(filename: str, data: bytes) -> list[dict]:
    """Return raw {header: value} rows from a CSV/XLSX upload. Raises ValueError."""
    ext = "." + (filename or "").rsplit(".", 1)[-1].lower() if "." in (filename or "") else ""
    if ext not in ALLOWED_EXTENSIONS:
        raise ValueError(f"Unsupported file type '{ext or '(none)'}'. Upload a .csv or .xlsx file.")
    if not data:
        raise ValueError("The file is empty.")
    if ext == ".csv":
        return _parse_csv(data)
    return _parse_xlsx(data)


def to_lead_fields(raw: dict) -> dict:
    """Map one raw row to canonical, parsed lead fields (only known columns)."""
    out: dict = {}
    for header, value in raw.items():
        field = _header_to_field(header)
        if not field:
            continue
        value = (str(value).strip() if value is not None else "")
        if value == "":
            continue
        if field == "lead_type":
            out[field] = _parse_lead_type(value)
        elif field == "source":
            out[field] = _parse_source(value)
        elif field == "team_size":
            out[field] = _parse_int(value)
        elif field in ("uses_software", "open_to_platform", "willing_for_demo"):
            out[field] = _parse_bool(value)
        elif field == "email":
            out[field] = value.lower()
        elif field == "phone":
            # Meta export prefixes phone with "p:" (e.g. "p:+919520771486").
            v = re.sub(r"^[A-Za-z]:", "", value).replace(" ", "").replace("-", "")
            out[field] = v
        else:
            out[field] = value
    return out


# ── Downloadable template ─────────────────────────────────────────────────────

TEMPLATE_HEADERS = [
    "Name", "Email", "Phone", "Company", "State", "Type",
    "Source", "Team Size", "Uses Software", "Open To Platform", "Wants Demo", "Website",
]
_TEMPLATE_SAMPLE = [
    ["Rajesh Sharma", "rajesh@brokerhouse.in", "9876543210", "Mehta Brokers",
     "Maharashtra", "broker", "referral", "12", "no", "yes", "yes", "mehtabrokers.in"],
    ["Priya Nair", "priya@agentpro.in", "9123456780", "AgentPro",
     "Karnataka", "agent", "referral", "3", "yes", "no", "no", ""],
]


def template_csv() -> str:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(TEMPLATE_HEADERS)
    w.writerows(_TEMPLATE_SAMPLE)
    return buf.getvalue()
